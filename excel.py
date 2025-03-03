from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, NamedStyle, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule

CELL_DIMENSION = 50
DAY_WIDTH = 20
team_row = {}

def create_sheet(rangliste: dict[str, list], mannschaften: set, club_teams: list, spiele: dict, runde: str, filename: str):
    workbook = Workbook()
    del workbook['Sheet']

    register_styles(workbook)
    sheet = workbook.create_sheet(runde)
    add_conditional_formatting(sheet, len(mannschaften) + 2)
    sheet.freeze_panes = sheet[f"D{len(mannschaften)+2}"] # D7

    damen = rangliste['damen']
    herren = rangliste['herren']


    length = schreibe_mannschaften(sheet, club_teams)
    schreibe_spiele(sheet, spiele)
    length = schreibe_rangliste_header(sheet, length)
    length = schreibe_rangliste_banner(sheet, "Damen", length)
    length = schreibe_rangliste(sheet, damen, length)
    length = schreibe_rangliste_banner(sheet, "Herren", length)
    length = schreibe_rangliste(sheet, herren, length)

    # workbook.save(filename=filename)
    return workbook

def add_conditional_formatting(sheet, index):
    text = Font(color="000000")
    red_fill = PatternFill(bgColor="FFC7CE")
    green_fill = PatternFill(bgColor="C6EFCE")
    neutral_fill = PatternFill(bgColor="FFEB9C")
    ja = DifferentialStyle(font=text, fill=green_fill)
    nein = DifferentialStyle(font=text, fill=red_fill)
    vielleicht = DifferentialStyle(font=text, fill=neutral_fill)

    ja_rule = Rule(type="containsText", operator="containsText", text="ja", dxf=ja)
    nein_rule = Rule(type="containsText", operator="containsText", text="nein", dxf=nein)
    vielleicht_rule = Rule(type="containsText", operator="containsText", text="vielleicht", dxf=vielleicht)

    sheet.conditional_formatting.add(f"D{index}:AZ200", ja_rule)
    sheet.conditional_formatting.add(f"D{index}:AZ200", nein_rule)
    sheet.conditional_formatting.add(f"D{index}:AZ200", vielleicht_rule)

def register_styles(workbook):
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side = Side(border_style="thin", color="31859B")
    border = Border(top=side, right=side)

    rangliste_header_style = NamedStyle(name='rangliste')
    rangliste_header_style.font = Font(bold=True, color="ffffff", size=12)
    rangliste_header_style.fill = PatternFill(fill_type="solid", fgColor="31859B")
    rangliste_header_style.alignment = center

    heimspiel = NamedStyle(name='heimspiel')
    heimspiel.fill = PatternFill(fill_type="solid", fgColor="FDE9D9")
    heimspiel.alignment = center
    # heimspiel.border = border

    gastspiel = NamedStyle(name='gastspiel')
    gastspiel.fill = PatternFill(fill_type="solid", fgColor="E5E0EC")
    gastspiel.alignment = center
    # gastspiel.border = border

    keinspiel = NamedStyle(name='keinspiel')
    keinspiel.border = border

    workbook.add_named_style(rangliste_header_style)
    workbook.add_named_style(heimspiel)
    workbook.add_named_style(gastspiel)
    workbook.add_named_style(keinspiel)

def schreibe_spiele(sheet, spieltermine: dict[str, list]):
    sheet.row_dimensions[1].height = 30
    column = 4
    for tag, spiele in spieltermine.items():
        if (len(spiele) == 0): continue

        sheet.cell(row=1, column=column, value=tag).style = 'rangliste'
        sheet.column_dimensions[get_column_letter(column)].width = DAY_WIDTH
        for spiel in spiele:
            value = f"Auswärts - {spiel['gegner']} {spiel['time']} Uhr"
            style = "gastspiel"
            if spiel['heim']:
                value = f"Heim - {spiel['gegner']} {spiel['time']} Uhr"
                style = "heimspiel"
            
            sheet.cell(row=team_row[spiel['mannschaft']], column=column, value=value).style = style
        column += 1

def schreibe_mannschaften(sheet, mannschaften, index=2):
    
    for i, mannschaft in enumerate(mannschaften, start=index):
        sheet.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        cell = sheet.cell(row=i, column=2, value=mannschaft)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        sheet.row_dimensions[i].height = CELL_DIMENSION
        team_row[mannschaft] = i

    return len(mannschaften) + index

def schreibe_rangliste(sheet, rangliste, index=0):
    max_length = 0
    for i, spieler in enumerate(rangliste, start=index):
        rang = spieler['rang']
        mannschaft = spieler['mannschaft']
        name = spieler['name']
        team = f"{mannschaft if len(mannschaft) < 3 else mannschaft[:3]}"

        if (len(name) > max_length):
            max_length = len(name) 
        
        sheet.row_dimensions[i].height = DAY_WIDTH
        sheet.cell(row=i, column=1, value=rang)
        sheet.cell(row=i, column=2, value=name)
        sheet.cell(row=i, column=3, value=team)
    
    sheet.column_dimensions['B'].width = max_length + 2

    return len(rangliste) + index

def schreibe_rangliste_header(sheet, index):
    sheet.cell(row=index, column=1, value="Rang").style = 'rangliste'
    sheet.cell(row=index, column=2, value="Name").style = 'rangliste'
    sheet.cell(row=index, column=3, value="Team").style = 'rangliste'

    return 1 + index

def schreibe_rangliste_banner(sheet, gender, index):
    sheet.merge_cells(start_row=index, start_column=1, end_row=index+2, end_column=3)
    cell = sheet.cell(row=index, column=1, value=gender)

    fill = PatternFill(fill_type="solid", fgColor="C2D69A")
    font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    cell.fill = fill
    cell.font = font
    cell.alignment = center

    return 3 + index

herren = [
    {
        "rang": "1",
        "mannschaft": "I",
        "name": "Pfeifer, Daniel"
    },
    {
        "rang": "2",
        "mannschaft": "I",
        "name": "Sonntag, Stefan"
    },
    {
        "rang": "3",
        "mannschaft": "I",
        "name": "Münster-Müller, Sascha"
    },
    {
        "rang": "4",
        "mannschaft": "I",
        "name": "Kuhfs, Patrick"
    },
    {
        "rang": "5",
        "mannschaft": "II",
        "name": "Damm, Simon"
    },
    {
        "rang": "6",
        "mannschaft": "II",
        "name": "Baumann, Malte"
    },
    {
        "rang": "7",
        "mannschaft": "II",
        "name": "Groborz, Albert"
    },
    {
        "rang": "8",
        "mannschaft": "Reservespieler",
        "name": "Haas, Joshua"
    },
    {
        "rang": "9",
        "mannschaft": "II",
        "name": "Lukas, Maurizio"
    },
    {
        "rang": "10",
        "mannschaft": "Reservespieler",
        "name": "Jung, Simon"
    },
    {
        "rang": "11",
        "mannschaft": "III",
        "name": "Wenzel, Sebastian"
    },
    {
        "rang": "12",
        "mannschaft": "III",
        "name": "Brüning, Michael"
    },
    {
        "rang": "13",
        "mannschaft": "III",
        "name": "Salis, Rene"
    },
    {
        "rang": "14",
        "mannschaft": "III",
        "name": "Buge, Jan"
    },
    {
        "rang": "15",
        "mannschaft": "IV",
        "name": "Wiedenmann, Dominik"
    },
    {
        "rang": "16",
        "mannschaft": "IV",
        "name": "Kisseler, Moritz"
    },
    {
        "rang": "17",
        "mannschaft": "Reservespieler",
        "name": "Walke, Lian"
    },
    {
        "rang": "18",
        "mannschaft": "Reservespieler",
        "name": "Yasni, Sebastian"
    },
    {
        "rang": "19",
        "mannschaft": "Reservespieler",
        "name": "Classen, Matthias"
    },
    {
        "rang": "20",
        "mannschaft": "IV",
        "name": "Nowacki, Jaro"
    },
    {
        "rang": "21",
        "mannschaft": "IV",
        "name": "Walke, Loris"
    },
    {
        "rang": "22",
        "mannschaft": "Reservespieler",
        "name": "Wagner, Fabian"
    },
    {
        "rang": "23",
        "mannschaft": "Reservespieler",
        "name": "Hofmann, Stefan"
    },
    {
        "rang": "24",
        "mannschaft": "Reservespieler",
        "name": "Wilhelm, Niklas"
    },
    {
        "rang": "25",
        "mannschaft": "Reservespieler",
        "name": "Schramm, Jannis"
    },
    {
        "rang": "26",
        "mannschaft": "Reservespieler",
        "name": "Schneider, Thomas"
    },
    {
        "rang": "27",
        "mannschaft": "Reservespieler",
        "name": "Sunny, Abhishek Mathew"
    },
    {
        "rang": "28",
        "mannschaft": "Reservespieler",
        "name": "Baier, Uwe"
    },
    {
        "rang": "29",
        "mannschaft": "Reservespieler",
        "name": "Nicklas, Luca Malte"
    },
    {
        "rang": "30",
        "mannschaft": "Reservespieler",
        "name": "Marter, Wolfgang"
    },
    {
        "rang": "31",
        "mannschaft": "Reservespieler",
        "name": "Siebert, Ralf"
    },
    {
        "rang": "32",
        "mannschaft": "Reservespieler",
        "name": "Holdermann, Thomas"
    },
    {
        "rang": "33",
        "mannschaft": "Reservespieler",
        "name": "Peter, Marcel"
    },
    {
        "rang": "34",
        "mannschaft": "Reservespieler",
        "name": "Malcher, Leon"
    },
    {
        "rang": "35",
        "mannschaft": "Reservespieler",
        "name": "Woznica, Marcel"
    },
    {
        "rang": "36",
        "mannschaft": "Reservespieler",
        "name": "Hopf, Michael"
    },
    {
        "rang": "37",
        "mannschaft": "Reservespieler",
        "name": "Walke, Daniel"
    }
]

damen = [{'rang': '1', 'mannschaft': 'I', 'name': 'Eschenbrenner, Nelly'}, {'rang': '2', 'mannschaft': 'I', 'name': 'Walke, Chira'}, {'rang': '3', 'mannschaft': 'I', 'name': 'Ille, Thea'}, {'rang': '4', 'mannschaft': 'II', 'name': 'Hümpfner, Magdalena'}, {'rang': '5', 'mannschaft': 'II', 'name': 'Weiland, Nia'}, {'rang': '6', 'mannschaft': 'II', 'name': 'Baumann, Miko'}, {'rang': '7', 'mannschaft': 'II', 'name': 'Kisseler, Cora'}, {'rang': '8', 'mannschaft': 'III', 'name': 'Paubandt, Beatrice'}, {'rang': '9', 'mannschaft': 'III', 'name': 'Siebert, Renée'}, {'rang': '10', 'mannschaft': 'Reservespieler', 'name': 'Ulrich, Laura'}, {'rang': '11', 'mannschaft': 'Reservespieler', 'name': 'Rodriguez-Feck, Emilia'}, {'rang': '12', 'mannschaft': 'IV', 'name': 'Scholz, Lisa'}, {'rang': '13', 'mannschaft': 'IV', 'name': 'Diemert, Sarah'}, {'rang': '14', 'mannschaft': 'IV', 'name': 'Hensel, Monika'}, {'rang': '15', 'mannschaft': 'Reservespieler', 'name': 'Kunig, Selin'}, {'rang': '16', 'mannschaft': 'Reservespieler', 'name': 'Marcelo, Khloe'}, {'rang': '17', 'mannschaft': 'Reservespieler', 'name': 'Hübner, Leandra'}]

rangliste = {"herren": herren, "damen": damen}
mannschaften = ["PSV GW Wiesbaden", "PSV GW Wiesbaden II", "PSV GW Wiesbaden III", "PSV GW Wiesbaden IV"]

spiele = {
    "01.09.2024": [
        {
            "time": "10:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "TV Wehen II"
        },
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "BC Ginsheim-Mainspitze V"
        },
        {
            "time": "13:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "SG Flörsheim/Rüsselsheim II"
        },
        {
            "time": "15:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "BC Ginsheim-Mainspitze IV"
        }
    ],
    "07.09.2024": [
        {
            "time": "17:30",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "TV 1904 Bermbach"
        }
    ],
    "21.09.2024": [
        {
            "time": "00:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "spielfrei"
        },
        {
            "time": "15:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "1. Wiesbadener BC III"
        },
        {
            "time": "15:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "TV Hofheim III"
        }
    ],
    "28.09.2024": [
        {
            "time": "15:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "TV Jahn Mensfelden"
        },
        {
            "time": "15:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "SG Flörsheim/Rüsselsheim"
        }
    ],
    "05.10.2024": [
        {
            "time": "15:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "BC Ginsheim-Mainspitze"
        },
        {
            "time": "16:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "TG Camberg IV"
        },
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "TV Idstein"
        }
    ],
    "02.11.2024": [
        {
            "time": "10:30",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "SG Flörsheim/Rüsselsheim III"
        }
    ],
    "03.11.2024": [
        {
            "time": "10:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "TV Idstein II"
        },
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "TuS Dotzheim"
        },
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "TG Camberg II"
        }
    ],
    "09.11.2024": [
        {
            "time": "15:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "SG TGH/TGU"
        },
        {
            "time": "19:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "TG Camberg"
        },
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "SV Fun-Ball Dortelweil IV"
        }
    ],
    "16.11.2024": [
        {
            "time": "15:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "SG Flörsheim/Rüsselsheim IV"
        },
        {
            "time": "17:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "TSV Bleidenstadt II"
        }
    ],
    "23.11.2024": [
        {
            "time": "17:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "1. Wiesbadener BC"
        },
        {
            "time": "18:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "TURA Niederhöchstadt"
        }
    ],
    "24.11.2024": [
        {
            "time": "00:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "spielfrei"
        },
        {
            "time": "09:30",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "Spvgg Hochheim"
        },
        {
            "time": "11:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "Rot-Weiss Walldorf"
        },
        {
            "time": "13:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "SG Eltville/Hallgarten"
        }
    ],
    "19.01.2025": [
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "BC Ginsheim-Mainspitze IV"
        },
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "SG Flörsheim/Rüsselsheim II"
        },
        {
            "time": "12:30",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "TV Wehen II"
        },
        {
            "time": "13:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "BC Ginsheim-Mainspitze V"
        },
        {
            "time": "19:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "1. Wiesbadener BC III"
        }
    ],
    "26.01.2025": [
        {
            "time": "00:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "spielfrei"
        },
        {
            "time": "11:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "SG Flörsheim/Rüsselsheim III"
        }
    ],
    "12.02.2025": [
        {
            "time": "19:30",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "TV Jahn Mensfelden"
        },
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "TV 1904 Bermbach"
        },
        {
            "time": "12:30",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "SG Eltville/Hallgarten"
        }
    ],
    "22.02.2025": [
        {
            "time": "15:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "TG Camberg IV"
        },
        {
            "time": "15:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "TV Hofheim III"
        },
        {
            "time": "18:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "TV Idstein"
        },
        {
            "time": "13:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "SG Flörsheim/Rüsselsheim"
        },
        {
            "time": "13:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "BC Ginsheim-Mainspitze"
        },
        {
            "time": "15:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "TV Idstein II"
        },
        {
            "time": "19:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "SG TGH/TGU"
        },
        {
            "time": "19:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "TG Camberg II"
        }
    ],
    "09.03.2025": [
        {
            "time": "10:30",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "TuS Dotzheim"
        },
        {
            "time": "11:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "SV Fun-Ball Dortelweil IV"
        }
    ],
    "15.03.2025": [
        {
            "time": "17:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "TG Camberg"
        },
        {
            "time": "17:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "SG Flörsheim/Rüsselsheim IV"
        }
    ],
    "16.03.2025": [
        {
            "time": "10:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "TSV Bleidenstadt II"
        }
    ],
    "22.03.2025": [
        {
            "time": "15:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "TURA Niederhöchstadt"
        },
        {
            "time": "18:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden II",
            "gegner": "1. Wiesbadener BC"
        }
    ],
    "23.03.2025": [
        {
            "time": "00:00",
            "heim": False,
            "mannschaft": "PSV GW Wiesbaden III",
            "gegner": "spielfrei"
        },
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden IV",
            "gegner": "Spvgg Hochheim"
        },
        {
            "time": "10:00",
            "heim": True,
            "mannschaft": "PSV GW Wiesbaden",
            "gegner": "Rot-Weiss Walldorf"
        }
    ]
}


# create_sheet(rangliste, mannschaften, spiele)